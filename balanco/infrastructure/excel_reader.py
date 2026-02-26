from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from ..domain.model import BalanceReport, LedgerSection, LineItem, ManualTotals, MonthRef, Money
from ..domain.ports import BalanceReader, SourceSpec
from .timer import step


# ============================================================
# LAYOUT (fixo, mas com guard-rails)
# ============================================================

@dataclass(frozen=True, slots=True)
class Layout:
    # Colunas: B..M
    COL_B: int = 2
    COL_C: int = 3
    COL_M: int = 13

    # Linhas fixas do layout informado
    ENTRADAS_NAME_ROW: int = 2
    ENTRADAS_HEADER_ROW: int = 2
    ENTRADAS_DATA_ROWS: Tuple[int, ...] = (3,)

    OUTRAS_NAME_ROW: int = 5
    OUTRAS_HEADER_ROW: int = 5
    OUTRAS_DATA_ROWS: Tuple[int, ...] = (6, 7)

    DESP_NAME_ROW: int = 9
    DESP_HEADER_ROW: int = 9
    DESP_START_ROW: int = 10

    TOTAL_GERAL_ANCHOR_LABEL: str = "total geral"  # na coluna B (âncora)

    AMAURILIO_ROW: int = 42
    ACOS_VITAL_ROW: int = 44

    # Segurança
    MAX_SCAN_ROWS: int = 2000
    MAX_CONSECUTIVE_BLANK: int = 8


# ============================================================
# NORMALIZAÇÃO DE COMPETÊNCIA
# ============================================================

_PT_MONTHS_FULL = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}
_PT_MONTHS_ABBR = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _normalize_header_to_ym(header: object, default_year: int = 2025) -> str | None:
    """
    Retorna competência canônica 'YYYY-MM' ou None.
    Suporta:
    - date/datetime (Excel frequentemente devolve isso)
    - 'YYYY-MM' e 'YYYY-MM-DD'
    - 'abril-25', 'março-25', etc
    - 'mar', 'abr', ... (abreviações)
    Regra: janeiro-26 => 2025-12 (dez/25 pago em jan/26)
    """
    if header is None:
        return None

    # 1) Data real (seu caso)
    if isinstance(header, datetime):
        y, m = header.year, header.month
        if y == 2026 and m == 1:
            return "2025-12"
        return f"{y:04d}-{m:02d}"

    if isinstance(header, date):
        y, m = header.year, header.month
        if y == 2026 and m == 1:
            return "2025-12"
        return f"{y:04d}-{m:02d}"

    # 2) String
    h = _norm(header)
    if not h or h == "total geral":
        return None

    # '2025-04' ou '2025-04-01'
    m_iso = re.match(r"^(\d{4})-(\d{2})(?:-(\d{2}))?$", h)
    if m_iso:
        y = int(m_iso.group(1))
        mm = int(m_iso.group(2))
        if y == 2026 and mm == 1:
            return "2025-12"
        return f"{y:04d}-{mm:02d}"

    # abrev: mar/abr/... (Despesas)
    if h in _PT_MONTHS_ABBR:
        mm = _PT_MONTHS_ABBR[h]
        return f"{default_year:04d}-{mm:02d}"

    # pt: "abril-25", "março-25"
    m_pt = re.match(r"^([a-zçãõáéíóú]+)\s*-\s*(\d{2})$", h)
    if m_pt:
        month_txt = m_pt.group(1)
        yy = int(m_pt.group(2))
        year = 2000 + yy
        mm = _PT_MONTHS_FULL.get(month_txt)
        if mm is None:
            return None
        if year == 2026 and mm == 1:
            return "2025-12"
        return f"{year:04d}-{mm:02d}"

    return None


# ============================================================
# PARSING DE VALORES
# ============================================================

def _to_decimal(v: Any) -> Decimal:
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))

    s = str(v).strip()
    if s == "":
        return Decimal("0")

    s = s.replace("\u00A0", " ").replace("R$", "").replace("$", "").strip()
    s = re.sub(r"[^\d\.\,\-]", "", s)
    if s in ("", "-", ".", ",", "-.", "-,"):
        return Decimal("0")

    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(".", "")

    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _money(v: Any) -> Money:
    return Money(_to_decimal(v))


def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def _is_blank_row(ws, row: int, col_start: int, col_end: int) -> bool:
    for c in range(col_start, col_end + 1):
        if _cell(ws, row, c) not in (None, ""):
            return False
    return True


# ============================================================
# FILE DISCOVERY
# ============================================================

def _find_workbook_file(logger: logging.Logger, folder: Path, name_hint: str) -> Path:
    if not folder.exists():
        raise FileNotFoundError(f"Diretório não encontrado: {folder}")

    candidates: List[Path] = []
    for ext in ("*.xlsx", "*.xlsm"):
        candidates.extend(folder.glob(ext))

    if not candidates:
        raise FileNotFoundError(f"Nenhum .xlsx/.xlsm encontrado em {folder}")

    hint = _norm(name_hint)
    matches = [p for p in candidates if hint in _norm(p.name) or hint in _norm(p.stem)]
    if not matches:
        matches = [p for p in candidates if ("balan" in _norm(p.name) and "2025" in _norm(p.name))]

    if not matches:
        logger.error("Arquivos disponíveis no diretório:")
        for p in sorted(candidates):
            logger.error(f" - {p.name}")
        raise FileNotFoundError(f"Arquivo com hint '{name_hint}' não encontrado em {folder}")

    matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    chosen = matches[0]
    logger.info(f"Arquivo selecionado: {chosen.name}")
    return chosen


# ============================================================
# LEITURA DE HEADERS / VALUES
# ============================================================

def _read_headers(ws, row: int, col_start: int, col_end: int) -> List[Any]:
    """
    ⚠️ IMPORTANTE:
    Retorna valores BRUTOS do Excel para preservar datetime/date.
    NÃO converta para str aqui.
    """
    out: List[Any] = []
    for c in range(col_start, col_end + 1):
        out.append(_cell(ws, row, c))
    return out


def _split_headers(headers: List[Any]) -> Tuple[List[Any], Optional[Any]]:
    if headers and _norm(headers[-1]) == "total geral":
        return headers[:-1], headers[-1]
    return headers, None


def _read_row_amounts_by_competence(
    ws,
    row: int,
    headers: List[Any],
    col_start: int,
    col_end: int,
    default_year: int = 2025,
) -> Tuple[Dict[str, Money], Optional[Money]]:
    total_excel: Optional[Money] = None
    bucket: Dict[str, Decimal] = {}

    idx = 0
    for c in range(col_start, col_end + 1):
        h = headers[idx]
        v = _cell(ws, row, c)
        idx += 1

        if _norm(h) == "total geral":
            if isinstance(v, str) and v.strip().startswith("="):
                total_excel = None
            else:
                total_excel = _money(v)
            continue

        ym = _normalize_header_to_ym(h, default_year=default_year)
        if ym is None:
            continue

        bucket[ym] = bucket.get(ym, Decimal("0")) + _to_decimal(v)

    return ({k: Money(v) for k, v in bucket.items()}, total_excel)


def _monthrefs_from_competences(competences: List[str]) -> List[MonthRef]:
    return [MonthRef(c) for c in competences]


# ============================================================
# READER
# ============================================================

class ExcelBalanceReader(BalanceReader):
    def __init__(self, logger: logging.Logger, layout: Layout | None = None, default_year: int = 2025):
        self.logger = logger
        self.layout = layout or Layout()
        self.default_year = default_year

    def read(self, spec: SourceSpec) -> BalanceReport:
        l = self.layout
        logger = self.logger

        base_dir = Path(spec.prefer_unc_path) if getattr(spec, "prefer_unc_path", None) else Path(spec.base_dir)

        with step(logger, "Verificar acesso ao diretório de rede"):
            if not base_dir.exists():
                raise FileNotFoundError(f"Diretório não existe: {base_dir}")
            try:
                _ = next(base_dir.iterdir(), None)
            except Exception as e:
                raise PermissionError(f"Sem permissão para listar diretório: {base_dir}. Detalhe: {e}")
            logger.info(f"Acesso OK ao diretório: {base_dir.resolve()}")

        with step(logger, "Localizar arquivo do balanço"):
            wb_path = _find_workbook_file(logger, base_dir, spec.workbook_name_hint)
            logger.info(f"Workbook path: {wb_path.resolve()}")

        with step(logger, "Abrir workbook (openpyxl)"):
            wb = load_workbook(filename=str(wb_path), data_only=False)
            logger.info(f"Workbook aberto. Abas: {wb.sheetnames}")

        with step(logger, "Selecionar aba alvo (apenas esta será lida)"):
            if spec.sheet_name not in wb.sheetnames:
                raise KeyError(f"Aba '{spec.sheet_name}' não encontrada. Abas: {wb.sheetnames}")
            ws = wb[spec.sheet_name]
            logger.info(f"Aba selecionada: {spec.sheet_name}")
            logger.info("Garantia: nenhuma outra aba será lida.")

        with step(logger, "Ler tabela: Entradas"):
            entradas = self._read_fixed_table(
                ws=ws,
                name_cell=(l.ENTRADAS_NAME_ROW, l.COL_B),
                header_row=l.ENTRADAS_HEADER_ROW,
                data_rows=list(l.ENTRADAS_DATA_ROWS),
                label_col=l.COL_B,
                col_start=l.COL_C,
                col_end=l.COL_M,
            )

        with step(logger, "Ler tabela: Outras Saídas/Investimento"):
            outras = self._read_fixed_table(
                ws=ws,
                name_cell=(l.OUTRAS_NAME_ROW, l.COL_B),
                header_row=l.OUTRAS_HEADER_ROW,
                data_rows=list(l.OUTRAS_DATA_ROWS),
                label_col=l.COL_B,
                col_start=l.COL_C,
                col_end=l.COL_M,
            )

        with step(logger, "Ler tabela: Despesas (varredura dinâmica)"):
            despesas = self._read_dynamic_table_until_anchor(
                ws=ws,
                name_cell=(l.DESP_NAME_ROW, l.COL_B),
                header_row=l.DESP_HEADER_ROW,
                start_row=l.DESP_START_ROW,
                anchor_label_norm=_norm(l.TOTAL_GERAL_ANCHOR_LABEL),
                label_col=l.COL_B,
                col_start=l.COL_C,
                col_end=l.COL_M,
            )

        with step(logger, "Ler Total Amaurilio (manual)"):
            amaurilio = self._read_manual_row(
                ws=ws,
                label_cell=(l.AMAURILIO_ROW, l.COL_B),
                header_row=l.ENTRADAS_HEADER_ROW,   # ✅ pega os meses da linha 2
                value_row=l.AMAURILIO_ROW,          # ✅ valores na 42
                col_start=l.COL_C,
                col_end=l.COL_M,
            )

        with step(logger, "Ler Total Aços Vital (manual)"):
            acos_vital = self._read_manual_row(
                ws=ws,
                label_cell=(l.ACOS_VITAL_ROW, l.COL_B),
                header_row=l.ENTRADAS_HEADER_ROW,   # ✅ pega os meses da linha 2
                value_row=l.ACOS_VITAL_ROW,         # ✅ valores na 44
                col_start=l.COL_C,
                col_end=l.COL_M,
            )

        manual = ManualTotals(amaurilio=amaurilio, acos_vital=acos_vital)

        logger.info("Leitura do Excel concluída. Construindo BalanceReport (domínio).")
        return BalanceReport(
            entradas=entradas,
            outras_saidas=outras,
            despesas=despesas,
            manual=manual,
            source_base_dir=str(base_dir),
            source_workbook_path=str(wb_path),
            source_sheet=spec.sheet_name,
        )

    def _read_fixed_table(
        self,
        ws,
        name_cell: Tuple[int, int],
        header_row: int,
        data_rows: List[int],
        label_col: int,
        col_start: int,
        col_end: int,
    ) -> LedgerSection:
        logger = self.logger

        name = str(_cell(ws, name_cell[0], name_cell[1]) or "").strip() or "Tabela"
        headers = _read_headers(ws, header_row, col_start, col_end)

        logger.info(f"[DEBUG] Headers brutos ({name}): {headers!r}")

        month_headers, _ = _split_headers(headers)
        competences: List[str] = []
        for h in month_headers:
            ym = _normalize_header_to_ym(h, default_year=self.default_year)
            logger.info(f"[DEBUG] header={h!r} -> ym={ym!r}")
            if ym:
                competences.append(ym)

        competences = sorted(set(competences))
        months = _monthrefs_from_competences(competences)

        logger.info(f"Tabela '{name}' (fixa). HeaderRow={header_row}. Competências={competences}")

        items: List[LineItem] = []
        for r in data_rows:
            label = str(_cell(ws, r, label_col) or "").strip() or "(sem label)"
            by_comp, total_excel = _read_row_amounts_by_competence(
                ws=ws,
                row=r,
                headers=headers,
                col_start=col_start,
                col_end=col_end,
                default_year=self.default_year,
            )

            by_month: Dict[MonthRef, Money] = {m: by_comp.get(m.raw, Money(Decimal("0"))) for m in months}

            items.append(LineItem(label=label, by_month=by_month, total_excel=total_excel))
            logger.info(f"Tabela '{name}' | Linha '{label}' lida (row={r}).")

        return LedgerSection(name=name, months=months, items=items)

    def _read_dynamic_table_until_anchor(
        self,
        ws,
        name_cell: Tuple[int, int],
        header_row: int,
        start_row: int,
        anchor_label_norm: str,
        label_col: int,
        col_start: int,
        col_end: int,
    ) -> LedgerSection:
        l = self.layout
        logger = self.logger

        name = str(_cell(ws, name_cell[0], name_cell[1]) or "").strip() or "Tabela"
        headers = _read_headers(ws, header_row, col_start, col_end)

        month_headers, _ = _split_headers(headers)
        competences: List[str] = []
        for h in month_headers:
            ym = _normalize_header_to_ym(h, default_year=self.default_year)
            if ym:
                competences.append(ym)

        competences = sorted(set(competences))
        months = _monthrefs_from_competences(competences)

        logger.info(f"Tabela '{name}' (dinâmica). HeaderRow={header_row}. Competências={competences}")

        items: List[LineItem] = []
        row = start_row
        blank_streak = 0

        while row <= l.MAX_SCAN_ROWS:
            bval = _norm(_cell(ws, row, label_col))

            if bval == anchor_label_norm:
                logger.info(f"Âncora '{anchor_label_norm}' encontrada na linha {row}. Parando leitura de '{name}'.")
                break

            if _is_blank_row(ws, row, label_col, col_end):
                blank_streak += 1
                if blank_streak >= l.MAX_CONSECUTIVE_BLANK:
                    logger.warning(f"{blank_streak} linhas vazias consecutivas. Parando leitura de '{name}' (row={row}).")
                    break
                row += 1
                continue

            blank_streak = 0
            label = str(_cell(ws, row, label_col) or "").strip() or "(sem label)"

            by_comp, total_excel = _read_row_amounts_by_competence(
                ws=ws,
                row=row,
                headers=headers,
                col_start=col_start,
                col_end=col_end,
                default_year=self.default_year,
            )

            by_month: Dict[MonthRef, Money] = {m: by_comp.get(m.raw, Money(Decimal("0"))) for m in months}
            items.append(LineItem(label=label, by_month=by_month, total_excel=total_excel))

            logger.info(f"Tabela '{name}' | Linha '{label}' lida (row={row}).")
            row += 1

        logger.info(f"Tabela '{name}' finalizada. Linhas lidas: {len(items)}")
        return LedgerSection(name=name, months=months, items=items)

    def _read_manual_row(
        self,
        ws,
        label_cell: Tuple[int, int],
        header_row: int,
        value_row: int,
        col_start: int,
        col_end: int,
    ) -> Dict[MonthRef, Money]:
        logger = self.logger
        label = str(_cell(ws, label_cell[0], label_cell[1]) or "").strip() or "(sem label)"

        # ✅ headers vêm da linha de cabeçalho (ex: 2)
        headers = _read_headers(ws, header_row, col_start, col_end)

        # manual: ignora "Total Geral" se existir e normaliza
        month_headers, _ = _split_headers(headers)

        competences: List[str] = []
        for h in month_headers:
            ym = _normalize_header_to_ym(h, default_year=self.default_year)
            if ym:
                competences.append(ym)

        # Mantém ordem (dedupe preservando ordem) — mais auditável que set+sorted
        seen = set()
        ordered: List[str] = []
        for c in competences:
            if c in seen:
                continue
            seen.add(c)
            ordered.append(c)

        months = _monthrefs_from_competences(ordered)

        # ✅ valores vêm da linha do total (42/44)
        by_comp, _total_excel = _read_row_amounts_by_competence(
            ws=ws,
            row=value_row,
            headers=headers,
            col_start=col_start,
            col_end=col_end,
            default_year=self.default_year,
        )

        by_month: Dict[MonthRef, Money] = {m: by_comp.get(m.key, Money(Decimal("0"))) for m in months}

        logger.info(
            f"Totais manuais '{label}' lidos (value_row={value_row}, header_row={header_row}). "
            f"Competências={ordered}"
        )
        return by_month